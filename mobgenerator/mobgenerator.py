import json
from dataclasses import dataclass
from enum import Enum

from openpyxl import load_workbook

from mobgenerator.domain import EnemyType, Enemy


class InvalidFileException(Exception):
    pass


class MetaKey(str, Enum):
    DELTA_TIME = "deltaTime"
    TRACKS = "tracks"
    INDEX = "index"


class EnemiesKey(str, Enum):
    ID = "ID"
    ARROW = "arrow"
    BULLET = "bullet"
    SPECIAL = "special"
    SPEED = "speed"
    RESOURCE_PATH = "resourcePath"


def _get_sheet(title, sheets):
    try:
        return [sheet for sheet in sheets if title in sheet.title][0]
    except IndexError:
        raise InvalidFileException(f"The spreadsheet does not contains '{title}' worksheet.")


def _try_get_property(prop_cell, val_cell, meta_key, initial_value=None):
    return val_cell.value if not initial_value and meta_key.value in prop_cell.value else initial_value


@dataclass
class MobGenerator:
    input_file_name: str
    delta_time: float = None
    tracks: int = None
    index: int = None

    def __post_init__(self):
        self._workbook = load_workbook(self.input_file_name)
        sheets = self._workbook.worksheets
        self._meta_sheet = _get_sheet("meta", sheets)
        self._load_from_meta_sheet()
        print("Meta loaded")
        self._enemies_sheet = _get_sheet("enemies", sheets)
        self._load_from_enemies_sheet()
        print("Enemies loaded")
        self._level_sheet = _get_sheet("level", sheets)
        self._load_from_level_sheet()
        print("Level loaded")

    def generate_file(self, output_file_name):
        file = open(f"outputs/{output_file_name}", "w")
        output_dict = {
            "index": self.index,
            "trackCount": self.tracks,
            "enemies": [enemy.get_representation() for enemy in self.enemies]
        }
        file.write(json.dumps(output_dict,indent=4))
        file.close()
        print("File generated successfully")

    def _load_from_meta_sheet(self):
        for prop in range(1, self._meta_sheet.max_row + 1):
            prop_cell = self._meta_sheet[f"A{prop}"]
            val_cell = self._meta_sheet[f"B{prop}"]
            self.delta_time = _try_get_property(prop_cell, val_cell, MetaKey.DELTA_TIME, self.delta_time)
            self.tracks = _try_get_property(prop_cell, val_cell, MetaKey.TRACKS, self.tracks)
            self.index = _try_get_property(prop_cell, val_cell, MetaKey.INDEX, self.index)
        if None in [self.delta_time, self.tracks, self.index]:
            raise InvalidFileException("Meta Sheet is invalid")

    def _load_from_enemies_sheet(self):
        start_char = ord("A")
        self._enemy_types = list()
        header_dict = dict()
        for char in range(0, self._enemies_sheet.max_column):
            current_char = chr(start_char + char)
            header_cell = self._enemies_sheet[f"{current_char}1"]
            header_dict[header_cell.value] = current_char
        for row in range(2, self._enemies_sheet.max_row + 1):
            try:
                id = self._enemies_sheet[f"{header_dict[EnemiesKey.ID.value]}{row}"].value
                arrow = self._enemies_sheet[f"{header_dict[EnemiesKey.ARROW.value]}{row}"].value
                bullet = self._enemies_sheet[f"{header_dict[EnemiesKey.BULLET.value]}{row}"].value
                special = self._enemies_sheet[f"{header_dict[EnemiesKey.SPECIAL.value]}{row}"].value
                speed = self._enemies_sheet[f"{header_dict[EnemiesKey.SPEED.value]}{row}"].value
                resource_path = self._enemies_sheet[f"{header_dict[EnemiesKey.RESOURCE_PATH.value]}{row}"].value
                self._enemy_types.append(
                    EnemyType(int(id), float(arrow), float(bullet), float(special), float(speed), resource_path))
            except TypeError:
                raise InvalidFileException("Invalid enemy sheet, check cell types")
            except KeyError:
                raise InvalidFileException("Invalid enemy sheet, check header")

    def _get_enemy_type(self, id):
        for type in self._enemy_types:
            if type.id == id:
                return type

    def _load_from_level_sheet(self):
        self.enemies = list()
        start_char = ord("A")
        for track in range(0, self.tracks):
            current_char = chr(start_char + track)
            for timestamp in range(1, self._level_sheet.max_row + 1):
                raw_enemy = self._level_sheet[f"{current_char}{timestamp}"].value
                if not raw_enemy:
                    continue
                self.enemies.append(
                    Enemy(self._get_enemy_type(int(raw_enemy)), round(float(timestamp) * self.delta_time, 2), track))
